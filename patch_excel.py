"""
Patch database_overview_combined_v2.xlsx:
  - Update headline numbers (13,494 / NL 8,356 / BE 5,138)
  - Fix country chart data block
  - Add DATA SOURCES section + bar chart at bottom
Saves to database_overview_combined_v3.xlsx (preserves v2 untouched).
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

INPUT  = Path('data/database_overview_combined_v2.xlsx')
OUTPUT = Path('data/database_overview_combined_v4.xlsx')

# ── Updated numbers ───────────────────────────────────────────────────────────
TOTAL = 13_494
NL    = 8_356
BE    = 5_138

SOURCES = [
    ('Overture Maps',  6_789, 3_854, 2_935),
    ('Foursquare',     2_627, 1_529, 1_098),
    ('OpenStreetMap',  2_358, 1_253, 1_105),
    ('TRACES (EU)',    1_720, 1_720,     0),
]

# ── Colour palette (same as original) ─────────────────────────────────────────
PRIMARY    = '1E3F20'
SAGE       = '7A9A7E'
AMBER      = 'D4A373'
BG_DARK    = '2B302A'
BG_LIGHT   = 'FAFAED'
SECTION_BG = 'E8F0E9'
TEXT       = '222222'
WHITE      = 'FFFFFF'
GRAY       = '888888'
LIGHT_GRAY = 'DEDEDE'

# ── Style helpers ─────────────────────────────────────────────────────────────
def px(c):   return PatternFill('solid', fgColor=c)
def fx(bold=False, size=11, color=TEXT, italic=False):
    return Font(name='Calibri', bold=bold, size=size, color=color, italic=italic)
def ax(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def col_w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w
def row_h(ws, row, h):
    ws.row_dimensions[row].height = h

def region_bg(ws, r1, c1, r2, c2, color):
    f = px(color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = f

def write_cell(ws, row, col, value, font=None, align=None):
    """Write value to a cell even if it's inside a merged region."""
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                c = ws.cell(row=mr.min_row, column=mr.min_col)
                break
    c.value = value
    if font:  c.font      = font
    if align: c.alignment = align

def merge_write(ws, r1, c1, r2, c2, value, fnt, bg=None, h='left', v='center'):
    # Only merge if not already merged
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    cell.font  = fnt
    cell.alignment = ax(h, v)
    if bg:
        region_bg(ws, r1, c1, r2, c2, bg)
    return cell

# ── Load workbook ─────────────────────────────────────────────────────────────
print(f'Loading {INPUT} ...')
wb = openpyxl.load_workbook(INPUT)
ws = wb['Analytics']

# =============================================================================
# 1. HEADLINE NUMBERS
# =============================================================================

# Header right: "Dataset: May 2026   |   Records: X"
write_cell(ws, 1, 11, f'Dataset: May 2026   |   Records: {TOTAL:,}')

# KPI card 1 – Total farms
write_cell(ws, 6, 2,  f'{TOTAL:,}')

# KPI card 2 – NL
write_cell(ws, 6, 6,  f'{NL:,}')
write_cell(ws, 8, 6,  f'Netherlands  ({NL/TOTAL*100:.1f}%)')

# KPI card 3 – BE
write_cell(ws, 6, 10, f'{BE:,}')
write_cell(ws, 8, 10, f'Belgium  ({BE/TOTAL*100:.1f}%)')

print('  Updated KPI cards and header.')

# =============================================================================
# 2. FIX COUNTRY CHART DATA BLOCK (col R=18, S=19, rows 13-15)
# =============================================================================
ws.cell(row=14, column=19).value = NL
ws.cell(row=15, column=19).value = BE

print('  Updated country chart data block.')

# =============================================================================
# 3. SOURCE DATA BLOCK in chart data columns (R=18, S=19, rows 23-28)
# =============================================================================
ws.cell(row=23, column=18).value = 'Source'
ws.cell(row=23, column=19).value = 'Farms'
ws.cell(row=23, column=18).font  = fx(size=9, color=GRAY)
ws.cell(row=23, column=19).font  = fx(size=9, color=GRAY)
for i, (src, tot, nl_c, be_c) in enumerate(SOURCES, 24):
    ws.cell(row=i, column=18).value = src
    ws.cell(row=i, column=19).value = tot
n_src = len(SOURCES)

print('  Written source data block to chart columns.')

# =============================================================================
# 4. NEW "DATA SOURCES" SECTION at rows 81-95
# =============================================================================
ROW_SEC   = 81   # section divider
ROW_HDR   = 82   # table header
ROW_DATA  = 83   # first data row  (83-86)
ROW_CHART = 87   # chart anchor row

# Extend background
region_bg(ws, ROW_SEC, 1, ROW_CHART + 20, 16, BG_LIGHT)

# Row heights
row_h(ws, ROW_SEC - 1, 12)   # spacer row 80
for r in range(ROW_SEC, ROW_CHART + 20):
    row_h(ws, r, 15)
row_h(ws, ROW_HDR, 18)

# Section divider
region_bg(ws, ROW_SEC, 1, ROW_SEC, 16, BG_LIGHT)
ws.merge_cells(start_row=ROW_SEC, start_column=2, end_row=ROW_SEC, end_column=16)
cell = ws.cell(row=ROW_SEC, column=2)
cell.value = '  DATA SOURCES'
cell.font  = fx(bold=True, size=10, color=WHITE)
cell.fill  = px(PRIMARY)
cell.alignment = ax('left', 'center')
row_h(ws, ROW_SEC, 20)

# Table header
hdr_cols = ['Source', 'Netherlands', 'Belgium', 'Total', 'Share']
for ci, h in enumerate(hdr_cols, 2):
    c = ws.cell(row=ROW_HDR, column=ci, value=h)
    c.font = fx(bold=True, color=WHITE, size=10)
    c.fill = px(SAGE)
    c.alignment = ax('center', 'center')

# Data rows
for ri, (src, tot, nl_c, be_c) in enumerate(SOURCES, ROW_DATA):
    alt = (ri - ROW_DATA) % 2 == 0
    bg  = SECTION_BG if alt else WHITE
    pct = f'{tot / TOTAL * 100:.1f}%'
    for ci, val in enumerate([src, nl_c, be_c, tot, pct], 2):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = px(bg)
        c.font = fx(size=10)
        c.alignment = ax('right' if ci > 2 else 'left', 'center')
        if isinstance(val, int):
            c.number_format = '#,##0'

# Total row
tot_row = ROW_DATA + n_src
region_bg(ws, tot_row, 2, tot_row, 6, BG_DARK)
for ci, val in enumerate(['TOTAL', NL, BE, TOTAL, '100.0%'], 2):
    c = ws.cell(row=tot_row, column=ci, value=val)
    c.font = fx(bold=True, color=WHITE, size=10)
    c.fill = px(BG_DARK)
    c.alignment = ax('right' if ci > 2 else 'left', 'center')
    if isinstance(val, int):
        c.number_format = '#,##0'
row_h(ws, tot_row, 18)

print('  Built DATA SOURCES section table.')

# =============================================================================
# 5. SOURCE BAR CHART  — horizontal, placed beside the table
# =============================================================================
ch = BarChart()
ch.type  = 'bar'
ch.title = 'Farms by Data Source'
ch.style = 10
ch.width  = 16
ch.height = 9
ch.legend = None
ch.x_axis.title = 'Number of Farms'

d = Reference(ws, min_col=19, min_row=23, max_row=23 + n_src)
k = Reference(ws, min_col=18, min_row=24, max_row=23 + n_src)
ch.add_data(d, titles_from_data=True)
ch.set_categories(k)

# Fix string category reference — pre-compute formula so f-string evaluates correctly
src_cat_formula = f"'Analytics'!$R$24:$R${23 + n_src}"
for ser in ch.series:
    ser.cat = AxDataSource(strRef=StrRef(f=src_cat_formula))

ch.series[0].graphicalProperties.solidFill      = AMBER
ch.series[0].graphicalProperties.line.solidFill = AMBER

dl = DataLabelList()
dl.showVal       = True
dl.showPercent   = False
dl.showCatName   = False
dl.showSerName   = False
dl.showLegendKey = False
dl.numFmt        = '#,##0'
ch.dataLabels = dl

# Place chart to the right of the source table (starts at col G = 7)
ws.add_chart(ch, f'G{ROW_HDR}')

print('  Added source bar chart.')

# =============================================================================
# 6. SAVE
# =============================================================================
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
size = OUTPUT.stat().st_size / 1024
print()
print('=' * 50)
print(f'  Saved : {OUTPUT}')
print(f'  Size  : {size:.0f} KB')
print(f'  Total farms updated: {TOTAL:,}')
print(f'  NL: {NL:,}  |  BE: {BE:,}')
print('=' * 50)
